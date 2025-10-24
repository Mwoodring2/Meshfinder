# app/dataio/import_excel.py
from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import re
import os
from pathlib import Path

# Prefer pandas if available (fast, robust); fall back to openpyxl
try:
    import pandas as _pd
    _HAS_PANDAS = True
except Exception:
    _pd = None
    _HAS_PANDAS = False

try:
    import openpyxl as _oxl
    _HAS_OPENPYXL = True
except Exception:
    _oxl = None
    _HAS_OPENPYXL = False

_CAN_READ = _HAS_PANDAS or _HAS_OPENPYXL

NORMALIZE = lambda s: re.sub(r"[^a-z0-9]+", "_", str(s).strip().lower()).strip("_")

# Canonical keys we want
REQUIRED_KEYS = {"name", "path"}
OPTIONAL_KEYS = {"project", "type", "tags"}

# Header synonyms map
SYNONYMS: Dict[str, List[str]] = {
    "name": [
        "name", "part_name", "part", "filename", "file_name", "asset", "label", "component", "parttitle"
    ],
    "path": [
        "path", "file_path", "filepath", "relative_path", "full_path", "file", "source", "location"
    ],
    "project": [
        "project", "project_", "project_no", "project_number", "project_num", "project#", "job", "sku", "id"
    ],
    "type": [
        "type", "category", "part_type", "class", "group"
    ],
    "tags": [
        "tags", "notes", "keywords", "comment", "comments", "meta"
    ],
}

def _best_sheet_pandas(xl) -> str:
    best = None; best_score = -1
    for name in xl.sheet_names:
        df = xl.parse(name)
        # score = non-empty cells in first 50 rows
        score = int(df.head(50).count().sum())
        if score > best_score:
            best, best_score = name, score
    return best or xl.sheet_names[0]

def _read_excel_table(xlsx_path: Path) -> Tuple[List[str], List[Dict[str, str]], Dict]:
    meta = {"engine": None, "sheet": None}
    if not _CAN_READ:
        raise RuntimeError("Neither pandas nor openpyxl is available")

    if _HAS_PANDAS:
        xl = _pd.ExcelFile(xlsx_path)
        sheet = _best_sheet_pandas(xl)
        df = xl.parse(sheet)
        meta.update({"engine": "pandas", "sheet": sheet})
        # Coerce all to str for safe processing; keep NaN as empty strings
        df = df.fillna("")
        headers = [str(c) for c in df.columns]
        rows = df.astype(str).to_dict(orient="records")
        return headers, rows, meta

    # Fallback: openpyxl minimalist reader
    wb = _oxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    # Choose the sheet with most non-empty cells in first 50 rows
    best = None; best_score = -1
    for ws in wb.worksheets:
        score = 0
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
            score += sum(1 for v in row if v not in (None, ""))
        if score > best_score:
            best, best_score = ws, score
    ws = best or wb.active
    meta.update({"engine": "openpyxl", "sheet": ws.title})
    # Extract header row (first non-empty row)
    headers = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if any(v not in (None, "") for v in row):
            headers = [str(v) if v is not None else "" for v in row]
            start_row = row
            break
    if not headers:
        return [], [], meta
    # Read remaining rows
    data_rows: List[Dict[str, str]] = []
    for row in ws.iter_rows(min_row=ws.min_row+1, values_only=True):
        d = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else ""
            d[str(h)] = "" if v is None else str(v)
        data_rows.append(d)
    return headers, data_rows, meta

def _invert_synonyms() -> Dict[str, str]:
    inv: Dict[str, str] = {}
    for canon, alts in SYNONYMS.items():
        inv[NORMALIZE(canon)] = canon
        for a in alts:
            inv[NORMALIZE(a)] = canon
    return inv

def _map_headers(headers: List[str]) -> Tuple[Dict[str, str], List[str]]:
    inv = _invert_synonyms()
    mapping: Dict[str, str] = {}
    unknown: List[str] = []
    for h in headers:
        key = NORMALIZE(h)
        canon = inv.get(key)
        if canon and canon not in mapping.values():
            mapping[h] = canon
        else:
            unknown.append(h)
    return mapping, unknown

@dataclass
class ImportRow:
    name: str
    path: str
    project: Optional[str] = None
    type: Optional[str] = None
    tags: Optional[str] = None

def import_parts_from_excel(xlsx_path: str | Path, base_dir: Optional[str | Path] = None):
    """
    Returns (rows: List[ImportRow], report: dict)
    - Robust to header variants
    - Resolves relative paths against base_dir (if provided)
    - Skips rows with missing REQUIRED keys; records reasons
    """
    xlsx_path = Path(xlsx_path)
    base = Path(base_dir) if base_dir else None
    headers, raw_rows, meta = _read_excel_table(xlsx_path)
    mapping, unknown = _map_headers(headers)

    rows: List[ImportRow] = []
    skipped: List[Dict[str, str]] = []

    # Build reverse header index for quick access
    for r in raw_rows:
        canon_row = {}
        for h, v in r.items():
            canon = mapping.get(h)
            if canon:
                canon_row[canon] = v.strip()
        # Validate required fields
        missing = [k for k in REQUIRED_KEYS if not canon_row.get(k)]
        if missing:
            skipped.append({"reason": f"missing {','.join(missing)}", "row": str(r)[:500]})
            continue

        # Resolve path if relative
        p = canon_row.get("path", "")
        p = p.replace("\\", "/")
        if base and not os.path.isabs(p):
            p_resolved = str((base / p).resolve())
        else:
            p_resolved = p

        rows.append(ImportRow(
            name=canon_row.get("name", ""),
            path=p_resolved,
            project=canon_row.get("project"),
            type=canon_row.get("type"),
            tags=canon_row.get("tags"),
        ))

    report = {
        "engine": meta.get("engine"),
        "sheet": meta.get("sheet"),
        "headers": headers,
        "mapped": mapping,
        "unknown_headers": unknown,
        "imported": len(rows),
        "skipped": len(skipped),
        "skip_samples": skipped[:5],
    }
    return rows, report
